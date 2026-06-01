from __future__ import annotations

import json
import re
import sys
import zipfile
from copy import copy
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


HEADERS = [
    "No",
    "Module",
    "Component",
    "Issue",
    "Priority",
    "Developer",
    "Status",
    "Comments",
    "Remark",
    "Tester",
    "Fixed Date",
    "Development",
    "Deployment",
    "Epic",
    "Sprint",
    "Release",
]

HEADER_ALIASES = {
    "no": "no",
    "module": "moduleName",
    "mainmodule": "moduleName",
    "component": "componentName",
    "submodule": "componentName",
    "navigation": "componentName",
    "issue": "title",
    "priority": "priority",
    "developer": "assignedToName",
    "dev": "assignedToName",
    "developmentowner": "assignedToName",
    "assignedto": "assignedToName",
    "status": "status",
    "comments": "comments",
    "remark": "remark",
    "tester": "testerAssignedToName",
    "testingowner": "testerAssignedToName",
    "testedby": "testerAssignedToName",
    "fixeddate": "fixedDate",
    "development": "developmentStatus",
    "deployement": "deploymentStatus",
    "deployment": "deploymentStatus",
    "epic": "epicTitle",
    "sprint": "sprintName",
    "release": "releaseName",
}

COLUMN_WIDTHS = {
    "A": 10,
    "B": 24,
    "C": 24,
    "D": 24,
    "E": 42,
    "F": 14,
    "G": 24,
    "H": 16,
    "I": 32,
    "J": 32,
    "K": 24,
    "L": 14,
    "M": 14,
    "N": 14,
    "O": 24,
    "P": 20,
    "Q": 20,
}

PRIORITY_STYLES = {
    "low": {
        "fill": PatternFill(fill_type="solid", fgColor="DCFCE7"),
        "font": Font(color="166534", bold=True),
    },
    "medium": {
        "fill": PatternFill(fill_type="solid", fgColor="FEF3C7"),
        "font": Font(color="92400E", bold=True),
    },
    "high": {
        "fill": PatternFill(fill_type="solid", fgColor="FED7AA"),
        "font": Font(color="9A3412", bold=True),
    },
    "critical": {
        "fill": PatternFill(fill_type="solid", fgColor="FECACA"),
        "font": Font(color="991B1B", bold=True),
    },
}

INVALID_SHEET_CHARACTERS = re.compile(r"[\[\]\*:/\\?]")


def normalize_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").strip().lower() if ch.isalnum())


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None

    normalized = str(value).strip()
    return normalized or None


def normalize_number(value: Any) -> int | None:
    if value in (None, ""):
        return None

    if isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    text_value = str(value).strip()
    if not text_value:
        return None

    try:
        return int(float(text_value))
    except ValueError:
        return None


def normalize_date(value: Any) -> str | None:
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return value.date().isoformat()

    if isinstance(value, date):
        return value.isoformat()

    text_value = str(value).strip()
    return text_value or None


def normalize_boolean(value: Any) -> bool | None:
    if value in (None, ""):
        return None

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return bool(value)

    normalized_value = str(value).strip().lower()

    if normalized_value in {"yes", "y", "true", "1", "done", "checked"}:
        return True

    if normalized_value in {"no", "n", "false", "0", "pending", "unchecked"}:
        return False

    return None


def normalize_priority_key(value: Any) -> str:
    normalized_value = str(value or "").strip().lower()
    return "".join(ch for ch in normalized_value if ch.isalnum())


def sanitize_sheet_title(value: Any, fallback: str) -> str:
    normalized = normalize_text(value) or fallback
    normalized = INVALID_SHEET_CHARACTERS.sub(" ", normalized)
    normalized = " ".join(normalized.split())

    if not normalized:
        normalized = fallback

    return normalized[:31]


def sanitize_file_name(value: Any, fallback: str) -> str:
    normalized = normalize_text(value) or fallback
    normalized = re.sub(r'[^A-Za-z0-9._ -]+', "-", normalized)
    normalized = normalized.strip(" .")

    if not normalized:
        normalized = fallback

    if not normalized.lower().endswith(".xlsx"):
        normalized = f"{normalized}.xlsx"

    return normalized


def normalize_export_sheets(raw_payload: Any) -> list[dict[str, Any]]:
    if isinstance(raw_payload, dict):
        raw_sheets = raw_payload.get("sheets")

        if not isinstance(raw_sheets, list):
            raise ValueError("Export payload must include a sheets array.")

        return [
            {
                "sheetName": sanitize_sheet_title(
                    raw_sheet.get("sheetName"), f"Issues {index + 1}"
                ),
                "rows": raw_sheet.get("rows") if isinstance(raw_sheet.get("rows"), list) else [],
            }
            for index, raw_sheet in enumerate(raw_sheets)
            if isinstance(raw_sheet, dict)
        ]

    if isinstance(raw_payload, list):
        return [
            {
                "sheetName": "Issues",
                "rows": raw_payload,
            }
        ]

    raise ValueError("Export payload must be a list of rows or an object with sheets.")


def build_workbook(raw_payload: Any) -> Workbook:
    sheets = normalize_export_sheets(raw_payload)
    workbook = Workbook()

    if workbook.active is not None:
        workbook.remove(workbook.active)

    if not sheets:
        sheets = [{"sheetName": "Issues", "rows": []}]

    for sheet in sheets:
        worksheet = workbook.create_sheet(title=sheet["sheetName"])
        worksheet.append(HEADERS)
        worksheet.freeze_panes = "A2"

        for row in sheet["rows"]:
            fixed_date = row.get("fixedDate")
            date_value: date | None = None

            if fixed_date:
                try:
                    date_value = datetime.fromisoformat(str(fixed_date)).date()
                except ValueError:
                    try:
                        date_value = date.fromisoformat(str(fixed_date)[:10])
                    except ValueError:
                        date_value = None

            worksheet.append(
                [
                    row.get("no"),
                    row.get("moduleName"),
                    row.get("componentName"),
                    row.get("title"),
                    row.get("priority"),
                    row.get("assignedToName"),
                    row.get("status"),
                    row.get("comments"),
                    row.get("remark"),
                    row.get("testerAssignedToName"),
                    date_value,
                    row.get("developmentStatus"),
                    row.get("deploymentStatus"),
                    row.get("epicTitle"),
                    row.get("sprintName"),
                    row.get("releaseName"),
                ]
            )

        for cell in worksheet[1]:
            cell_font = copy(cell.font)
            cell_font.bold = True
            cell.font = cell_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2):
            priority_cell = row[4]
            priority_key = normalize_priority_key(priority_cell.value)
            priority_style = PRIORITY_STYLES.get(priority_key)

            if priority_style:
                priority_cell.fill = priority_style["fill"]
                priority_cell.font = priority_style["font"]

            priority_cell.alignment = Alignment(horizontal="center", vertical="center")
            row[10].alignment = Alignment(horizontal="center", vertical="center")
            row[11].alignment = Alignment(horizontal="center", vertical="center")
            row[12].alignment = Alignment(horizontal="center", vertical="center")

            if row[10].value:
                row[10].number_format = "yyyy-mm-dd"

        for column_name, width in COLUMN_WIDTHS.items():
            worksheet.column_dimensions[column_name].width = width

    return workbook


def export_rows(input_path: Path, output_path: Path) -> None:
    raw_payload = json.loads(input_path.read_text(encoding="utf-8-sig"))
    workbook = build_workbook(raw_payload)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def export_bundle(input_path: Path, output_path: Path) -> None:
    raw_payload = json.loads(input_path.read_text(encoding="utf-8-sig"))
    raw_workbooks = raw_payload.get("workbooks") if isinstance(raw_payload, dict) else None

    if not isinstance(raw_workbooks, list) or len(raw_workbooks) == 0:
        raise ValueError("Export bundle payload must include at least one workbook.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for index, raw_workbook in enumerate(raw_workbooks):
            if not isinstance(raw_workbook, dict):
                continue

            file_name = sanitize_file_name(
                raw_workbook.get("fileName"), f"issues-export-{index + 1}.xlsx"
            )
            workbook = build_workbook({"sheets": raw_workbook.get("sheets", [])})
            buffer = BytesIO()
            workbook.save(buffer)
            archive.writestr(file_name, buffer.getvalue())


def import_rows(input_path: Path, output_path: Path) -> None:
    workbook = load_workbook(input_path, data_only=True)
    parsed_sheets = []

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))

        if not rows:
            continue

        header_row = rows[0]
        if all(cell in (None, "") for cell in header_row):
            continue

        canonical_headers = []

        for header in header_row:
            header_key = normalize_header(header)
            canonical_header = HEADER_ALIASES.get(header_key)

            if canonical_header is None:
                canonical_headers.append(None)
                continue

            canonical_headers.append(canonical_header)

        missing_headers = [
            header
            for header_key, header in (
                ("no", "No"),
                ("moduleName", "Module"),
                ("componentName", "Component"),
                ("title", "Issue"),
                ("priority", "Priority"),
                ("assignedToName", "Developer"),
                ("status", "Status"),
                ("comments", "Comments"),
                ("remark", "Remark"),
                ("testerAssignedToName", "Tester"),
                ("fixedDate", "Fixed Date"),
                ("developmentStatus", "Development"),
                ("deploymentStatus", "Deployment"),
                ("epicTitle", "Epic"),
                ("sprintName", "Sprint"),
                ("releaseName", "Release"),
            )
            if header_key not in canonical_headers
        ]

        if missing_headers:
            raise ValueError(
                f'Sheet "{worksheet.title}" is missing required Excel header(s): {", ".join(missing_headers)}'
            )

        parsed_rows = []

        for row_number, values in enumerate(rows[1:], start=2):
            row_payload: dict[str, Any] = {
                "rowNumber": row_number,
                "no": None,
                "moduleName": None,
                "componentName": None,
                "title": None,
                "priority": None,
                "assignedToName": None,
                "testerAssignedToName": None,
                "status": None,
                "comments": None,
                "remark": None,
                "fixedDate": None,
                "developmentStatus": None,
                "deploymentStatus": None,
                "epicTitle": None,
                "sprintName": None,
                "releaseName": None,
            }

            is_empty = True

            for header_key, cell_value in zip(canonical_headers, values):
                if header_key is None:
                    continue

                if cell_value not in (None, ""):
                    is_empty = False

                if header_key == "no":
                    row_payload["no"] = normalize_number(cell_value)
                elif header_key == "fixedDate":
                    row_payload["fixedDate"] = normalize_date(cell_value)
                elif header_key in {"developmentStatus", "deploymentStatus"}:
                    row_payload[header_key] = normalize_text(cell_value)
                else:
                    row_payload[header_key] = normalize_text(cell_value)

            if is_empty:
                continue

            parsed_rows.append(row_payload)

        parsed_sheets.append({"sheetName": worksheet.title, "rows": parsed_rows})

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps({"sheets": parsed_sheets}), encoding="utf-8")


def main() -> int:
    if len(sys.argv) != 4:
        raise SystemExit(
            "Usage: issues_excel.py <export|export-bundle|import> <input> <output>"
        )

    command = sys.argv[1]
    input_path = Path(sys.argv[2])
    output_path = Path(sys.argv[3])

    if command == "export":
        export_rows(input_path, output_path)
        return 0

    if command == "export-bundle":
        export_bundle(input_path, output_path)
        return 0

    if command == "import":
        import_rows(input_path, output_path)
        return 0

    raise SystemExit(f"Unsupported command: {command}")


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:  # pragma: no cover - CLI failure path
        print(str(error), file=sys.stderr)
        raise SystemExit(1)
